from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
SRC = ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from nametag_generator.config import GeneratorConfig
from nametag_generator.models import Attendee
from nametag_generator.service import generate_pdf


INPUT_XLSX_CANDIDATES = [
    ROOT / "OpenAI Codex Meetup - Seoul.xlsx",
    Path("/Users/seowoohan/Desktop/OpenAI Codex Meetup - Seoul.xlsx"),
]
OUTPUT_PDF = ROOT / "output/pdf/openai-codex-meetup-seoul-name-tags.pdf"

ROLE_MAP = {
    "한서우": "STAFF",
    "이지영": "STAFF",
    "송승인": "STAFF",
    "황현태": "SPEAKER",
    "정구봉": "SPEAKER",
    "허예찬": "SPEAKER",
}


def resolve_input_xlsx() -> Path:
    for candidate in INPUT_XLSX_CANDIDATES:
        if candidate.exists():
            return candidate
    raise FileNotFoundError("OpenAI Codex Meetup - Seoul.xlsx not found")


def load_legacy_attendees(path: Path) -> list[Attendee]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook["참석자"] if "참석자" in workbook.sheetnames else workbook.active
    attendees: list[Attendee] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        attendees.append(
            Attendee(
                name=str(row[1]).strip(),
                organization=str(row[2]).strip() if row[2] else None,
                role=ROLE_MAP.get(str(row[1]).strip()),
                title=str(row[3]).strip() if row[3] else None,
            )
        )
    return attendees


def main() -> None:
    updated_attendees = load_legacy_attendees(resolve_input_xlsx())
    result = generate_pdf(
        updated_attendees,
        config=GeneratorConfig(),
        output_path=OUTPUT_PDF,
    )
    print(f"created {OUTPUT_PDF}")
    print(f"attendees {len(updated_attendees)}")
    print(f"font {result.font_path}")


if __name__ == "__main__":
    main()
