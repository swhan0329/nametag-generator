from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from nametag_generator.errors import WorkbookValidationError
from nametag_generator.models import Attendee


HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "name": ("name", "이름"),
    "organization": ("organization", "org", "소속"),
    "role": ("role", "역할", "구분", "참여자구분"),
    "title": ("title", "subtitle", "job_title", "직무"),
}

REQUIRED_FIELDS: tuple[str, ...] = ("name", "organization")


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "_")


def _build_column_map(headers: list[object]) -> dict[str, int]:
    normalized_headers = [_normalize_header(header) for header in headers]
    resolved: dict[str, int] = {}
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            normalized_alias = alias.strip().lower().replace(" ", "_")
            if normalized_alias in normalized_headers:
                resolved[field] = normalized_headers.index(normalized_alias)
                break
    missing = [field for field in REQUIRED_FIELDS if field not in resolved]
    if missing:
        missing_text = ", ".join(missing)
        raise WorkbookValidationError(
            f"Workbook is missing required columns: {missing_text}"
        )
    return resolved


def load_attendees_from_xlsx(
    path: Path,
    sheet_name: str | None = None,
) -> list[Attendee]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise WorkbookValidationError("Workbook is empty.")
    column_map = _build_column_map(list(rows[0]))

    attendees: list[Attendee] = []
    for row in rows[1:]:
        name_cell = row[column_map["name"]] if column_map["name"] < len(row) else None
        if name_cell is None or not str(name_cell).strip():
            continue
        organization_cell = (
            row[column_map["organization"]]
            if column_map["organization"] < len(row)
            else None
        )
        role_index = column_map.get("role")
        role_cell = row[role_index] if role_index is not None and role_index < len(row) else None
        title_index = column_map.get("title")
        title_cell = (
            row[title_index] if title_index is not None and title_index < len(row) else None
        )
        attendees.append(
            Attendee(
                name=str(name_cell).strip(),
                organization=str(organization_cell).strip()
                if organization_cell
                else None,
                role=str(role_cell).strip() if role_cell else None,
                title=str(title_cell).strip() if title_cell else None,
            )
        )

    if not attendees:
        raise WorkbookValidationError("Workbook does not contain any attendee rows.")
    return attendees
